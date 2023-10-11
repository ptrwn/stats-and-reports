from openpyxl.chart import (
    DoughnutChart,
    BarChart,
    Reference
)
import pandas as pd
from typing import Dict
from openpyxl import Workbook
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows


def draw_csat_doughnut(wb: Workbook, **csat_data: Dict) -> Workbook:
    '''Creates a sheet with doughnut chart in an existing book.
    
    Args:
        wb: workbook, a sheet with chart will be added there
        csat_data: summary of stats to visualize, produced by stat_counter 

    Returns:
        wb: updated workbook
    '''

    export_list = [
        ['category', 'inner', 'outer'],
        ['Irrelevant', csat_data['Irrelevant'], float('nan')],
        ['Relevant', csat_data['Relevant'], float('nan')],
        ['Mixed', float('nan'), csat_data['Mixed']],
        ['Duplicate', float('nan'), csat_data['Duplicate']],
        ['VSAT', float('nan'), csat_data['VSAT']],
        ['SAT', float('nan'), csat_data['SAT']],
        ['Neutral', float('nan'), csat_data['Neutral']],
        ['DSAT', float('nan'), csat_data['DSAT']],
        ['VDSAT', float('nan'), csat_data['VDSAT']]
    ]

    ws = wb.create_sheet('csat_doughnut')

    for row in export_list:
        ws.append(row)

    chart = DoughnutChart()
    labels = Reference(ws, min_col=1, min_row=2, max_row=10)
    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=10)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.style = 2
    chart.title = "CSAT = {}%".format(csat_data['CSAT score']*100)

    slices = [DataPoint(idx=i) for i in range(9)]
    irrel, rel, mixed, dup, vsat, sat, neu, dsat, vdsat = slices
    chart.series[0].data_points = slices
    chart.series[1].data_points = slices

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showPercent = True

    irrel.graphicalProperties.pattFill = PatternFillProperties(prst="pct20")
    rel.graphicalProperties.solidFill = "0080FF"
    mixed.graphicalProperties.pattFill = PatternFillProperties(prst="ltVert")
    dup.graphicalProperties.pattFill = PatternFillProperties(prst="ltHorz")
    vsat.graphicalProperties.solidFill = "008800"
    sat.graphicalProperties.solidFill = "00BB00"
    neu.graphicalProperties.solidFill = "FFFF00"
    dsat.graphicalProperties.solidFill = "FFAA00"
    vdsat.graphicalProperties.solidFill = "FF0000"

    ws.add_chart(chart, "E2")
    return wb 


def draw_dsat_reason_bars(wb: Workbook, dsats: pd.DataFrame) -> Workbook:
    '''Creates a sheet with bar chart in an existing book.
    
    Args:
        wb: workbook, a sheet with chart will be added there
        dsats: summary of dsat reasons to visualize, produced by get_dsat_reasons 

    Returns:
        wb: updated workbook
    '''

    num_reas = len(dsats)
    ws = wb.create_sheet('dsat_reason_bars')
    for row in dataframe_to_rows(dsats, index=False, header=True):
        ws.append(row)

    chart = BarChart()
    chart.type = 'col'
    chart.style = 10
    chart.y_axis.title = 'count'
    chart.x_axis.title = 'dsat drivers'
    titles = Reference(ws, min_col=1, max_col=1, min_row=2, max_row=num_reas + 1)
    data = Reference(ws, min_col=2, max_col=2, min_row=1, max_row=num_reas + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(titles)
    chart.shape = 4
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    chart.title = "Dissatisfaction reasons"
    ws.add_chart(chart, "E2")

    return wb
